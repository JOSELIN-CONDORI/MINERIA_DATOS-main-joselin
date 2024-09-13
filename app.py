from flask import Flask, render_template, request, redirect, url_for, send_file
from flask_login import LoginManager, login_required, current_user
from auth import auth, User, crear_conexion
from scraping_script import obtener_articulos_el_comercio, obtener_articulos_diario_sin_fronteras, obtener_articulos_trome
from db_logging import registrar_scraping, obtener_reportes_scraping
import mysql.connector
from mysql.connector import Error
import pandas as pd
import os
from datetime import datetime
from io import BytesIO
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import openpyxl  # Asegúrate de importar openpyxl


app = Flask(__name__)
app.secret_key = 'tu_clave_secreta'

# Configurar Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'auth.login'

# Registrar el blueprint de auth
app.register_blueprint(auth)

# Cargar el usuario
@login_manager.user_loader
def load_user(user_id):
    conexion = crear_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
    user = cursor.fetchone()
    cursor.close()
    conexion.close()

    if user:
        return User(id=user['id'], username=user['username'])
    return None

# Crear la carpeta "descargas" si no existe
if not os.path.exists('descargas'):
    os.makedirs('descargas')

# Función para guardar los artículos en la base de datos sin duplicados
def guardar_articulos(articulos):
    conexion = crear_conexion()
    if conexion is None:
        return

    try:
        cursor = conexion.cursor()
        for articulo in articulos:
            cursor.execute("SELECT id FROM articulos WHERE url_articulo = %s", (articulo['url_articulo'],))
            resultado = cursor.fetchone()

            if resultado:
                print(f"El artículo '{articulo['titulo']}' ya existe en la base de datos.")
            else:
                cursor.execute("""
                    INSERT INTO articulos (diario, titulo, url_articulo, imagen, contenido, fecha, autor)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    articulo['diario'],
                    articulo['titulo'],
                    articulo['url_articulo'],
                    articulo['imagen'],
                    articulo['contenido'],
                    articulo['fecha'],
                    articulo['autor']
                ))
                print(f"Artículo '{articulo['titulo']}' guardado correctamente en la base de datos.")
        
        conexion.commit()
    except Error as e:
        print(f"Error al guardar los artículos: {e}")
    finally:
        if conexion.is_connected():
            cursor.close()
            conexion.close()

def exportar_articulos_a_excel(articulos, diario):
    try:
        # Crear un DataFrame con los artículos que se están mostrando en la pantalla
        df = pd.DataFrame(articulos)

        # Crear un archivo Excel en memoria
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Articulos')
        writer.close()

        # Cargar el archivo de openpyxl para aplicar estilos
        output.seek(0)
        wb = openpyxl.load_workbook(output)
        ws = wb.active

        # Estilo para encabezados: Negrita
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Ajuste automático de ancho de columna
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            max_length = min(length, 50)  # Limitar el ancho máximo a 50 caracteres
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length

        # Guardar el archivo en la carpeta 'descargas'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'articulos_{diario}_{timestamp}.xlsx'.replace(" ", "_")
        ruta_archivo = os.path.join('descargas', nombre_archivo)

        # Guardar el archivo localmente
        wb.save(ruta_archivo)
        print(f"Archivo guardado en {ruta_archivo}")

        # Enviar el archivo Excel generado al navegador
        output.seek(0)  # Resetear el puntero al inicio para enviarlo
        output = BytesIO()  # Reiniciar el buffer para guardarlo de nuevo
        wb.save(output)
        output.seek(0)
        
        return send_file(output, as_attachment=True, download_name=nombre_archivo, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Error al exportar los artículos a Excel: {e}")
        return None

@app.route('/reportes')
@login_required
def reportes():
    reportes = obtener_reportes_scraping()
    return render_template('reportes.html', reportes=reportes)

# Ruta de scraping con paginación
@app.route('/admin/scraping')
@login_required
def scraping_page():
    url = request.args.get('url')
    page = int(request.args.get('page', 1))
    articulos = []
    diario = ''

    if url:
        try:
            # Identificar el diario según la URL y registrar el scraping
            if "elcomercio.pe" in url:
                articulos = obtener_articulos_el_comercio(url)
                diario = 'El Comercio'
            elif "diariosinfronteras.com.pe" in url:
                articulos = obtener_articulos_diario_sin_fronteras(url)
                diario = 'Diario Sin Fronteras'
            elif "trome.pe" in url:
                articulos = obtener_articulos_trome(url)
                diario = 'Trome'
            else:
                articulos = []
                print(f"URL no compatible: {url}")
            
            # Registrar el scraping en la base de datos
            registrar_scraping(diario)

        except Exception as e:
            print(f"Error al procesar la URL: {url}. Error: {e}")
            articulos = []

        if articulos:
            guardar_articulos(articulos)

    total_articulos = len(articulos)
    per_page = 10
    total_pages = (total_articulos // per_page) + (1 if total_articulos % per_page > 0 else 0)
    start = (page - 1) * per_page
    end = start + per_page

    return render_template('scraping.html', articulos=articulos[start:end], diario=diario, page=page, total_pages=total_pages)



    

# Ruta para descargar el Excel
@app.route('/descargar_excel', methods=['POST'])
@login_required
def descargar_excel():
    try:
        # Obtener los artículos desde el formulario en formato JSON
        articulos_json = request.form.get('articulos')
        diario = request.form.get('diario')  # Obtener el nombre del diario
        articulos = json.loads(articulos_json)
    except json.JSONDecodeError as e:
        print(f"Error al decodificar JSON: {e}")
        return "Error al procesar los datos.", 400

    return exportar_articulos_a_excel(articulos, diario)

# Ruta principal (redirige a login si no autenticado)
@app.route('/')
def home():
    return redirect(url_for('auth.login'))

if __name__ == '__main__':
    app.run(debug=True)
